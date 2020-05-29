from openpyxl import load_workbook
from src.util.sqlutil import SqlUtil
import re



class ExcelUtil(object):

    def get_stock_code(self, excel_name):
        wb = load_workbook(excel_name)

        ws = wb['stock_all']

        rows = ws.max_row  # 获取行数
        cols = ws.max_column  # 获取列数

        # for i in range(1, rows + 1):
        #    for j in range(1, cols + 1):
        #        print(ws.cell(row = i, column = j).value)
        #pattern = re.compile(r'[0-9]+')

        sqlUtil = SqlUtil()



        for i in range(2, rows + 1):

            matchObj = re.search(r'[0-9]+', ws.cell(row=i, column=2).value, re.M | re.I)
            if matchObj:
                code = matchObj.group()
            else:
                code = ''



            name = ws.cell(row=i, column=3).value.strip()

            industry = ws.cell(row=i, column=15).value.strip()

            print(code)
            print(name)
            print(industry)

            sqlUtil.insert_stock2(code, name, industry, '', '', '');


    def get_stock_concept(self, excel_name):
        wb = load_workbook(excel_name)

        ws = wb['stock_concept_apple']

        rows = ws.max_row  # 获取行数
        cols = ws.max_column  # 获取列数

        # for i in range(1, rows + 1):
        #    for j in range(1, cols + 1):
        #        print(ws.cell(row = i, column = j).value)
        #pattern = re.compile(r'[0-9]+')

        sqlUtil = SqlUtil()



        for i in range(2, rows + 1):

            concept = ws.cell(row=i, column=2).value.strip()

            matchObj = re.search(r'[0-9]+', ws.cell(row=i, column=3).value, re.M | re.I)
            if matchObj:
                code = matchObj.group()
            else:
                code = ''



            name = ws.cell(row=i, column=4).value.strip()

            industry = ws.cell(row=i, column=5).value.strip()

            print(concept)
            print(code)
            print(name)
            print(industry)

            sqlUtil.insert_stock_concept(concept, code, name);



    def get_stock_industry(self, excel_name):

        wb = load_workbook(excel_name)

        ws = wb['industry']

        rows = ws.max_row  # 获取行数
        cols = ws.max_column  # 获取列数

        # for i in range(1, rows + 1):
        #    for j in range(1, cols + 1):
        #        print(ws.cell(row = i, column = j).value)
        #pattern = re.compile(r'[0-9]+')

        sqlUtil = SqlUtil()



        for i in range(2, rows + 1):

            code = ws.cell(row=i, column=1).value
            name = ws.cell(row=i, column=2).value

            industry = ws.cell(row=i, column=3).value
            segment = ws.cell(row=i, column=4).value


            print(code[2:len(code)])
            print(code)
            print(name)
            print(industry)
            print(segment)

            sqlUtil.insert_stock_industry(code[2:len(code)], name, industry, segment);






if __name__ == "__main__":

    excelUtil = ExcelUtil()
    excel_name = '/Users/xianxiaoge/PycharmProjects/stock/data/industry/stock_all.xlsx'
    excelUtil.get_stock_industry(excel_name)




